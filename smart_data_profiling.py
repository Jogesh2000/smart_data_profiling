# -*- coding: utf-8 -*-
"""
Created on Jan 6 23:36:10 2025

@author: Jogesh
"""

import pandas as pd
from ydata_profiling import ProfileReport
import pyodbc
import re
from openpyxl import Workbook

# Path to the Excel configuration file
config_file_path = "C:/py/profiling_report/dg_profiling_config_file.xlsx"

# Read the Excel file
sheet_data = pd.read_excel(config_file_path, header=None)

# Find the last column before "END OF DATASETS" in Row 1
end_marker = "END OF DATASETS"
last_column = sheet_data.iloc[0].eq(end_marker).idxmax()

# Extract relevant columns for processing
relevant_columns = sheet_data.iloc[:, 1:last_column]  # Ignore Column A

# Initialize a dictionary to store table names, domains, and their attributes
tables_and_attributes = {}

# Extract datasets and attributes for each relevant column
for col in relevant_columns.columns:
    table_name = sheet_data.iloc[5, col]  # Table name in row 5
    domain_name = sheet_data.iloc[2, col]  # Domain name in row 2

    if pd.isna(table_name) or pd.isna(domain_name):
        continue  # Skip if table name or domain name is not defined

    # Extract attributes and aliases for the current table starting from row 6
    attributes = []
    aliases = []
    for value in sheet_data.iloc[6:, col].dropna():
        # Extract text before the first space followed by `(` if it exists, else take the whole value
        if " (" in str(value):
            attr = re.split(r' \(', str(value))[0]
        else:
            attr = str(value)
        attributes.append(attr)
        aliases.append(value)  # Use the full cell value as the alias

    tables_and_attributes[(table_name, domain_name)] = {"attributes": attributes, "aliases": aliases}

# Database connection details (replace with actual values)
conn = pyodbc.connect(
    'Driver={SQL Server};'
    'Server=your_server;'
    'Database=your_database;'
    'uid=your_user;'
    'pwd=your_password;'
    'Trusted_Connection=no;'
)

# Initialize file paths
profile_report_html_files = []
data_quality_exception_files = []

# Data quality checks
def null_or_empty_check(df, aliases):
    exceptions = []
    for alias in aliases:
        null_values = df[df[alias].isnull() | (df[alias] == "")][alias].dropna().unique().tolist()
        if null_values:
            exceptions.append((alias, "Null or Empty Check", null_values))
    return exceptions

def frequent_category_check(df, aliases):
    exceptions = []
    for alias in aliases:
        most_common_percentage = df[alias].value_counts(normalize=True).max()
        if most_common_percentage > 0.8:  # Example: Flag if any value exceeds 80% of the total
            most_common_value = [df[alias].value_counts().idxmax()]
            exceptions.append((alias, "Frequent Category Check", most_common_value))
    return exceptions

def length_consistency_check(df, alias, min_length, max_length):
    invalid_lengths = df[(df[alias].str.len() < min_length) | (df[alias].str.len() > max_length)][alias].dropna().unique().tolist()
    if invalid_lengths:
        return [(alias, "Length Consistency Check", invalid_lengths)]
    return []

def product_data_quality_checks(df, aliases):
    exceptions = []
    for alias in aliases:
        lengths = df[alias].astype(str).str.len()
        median_length = lengths.median()
        outlier_indices = lengths[(lengths < 0.8 * median_length) | (lengths > 1.2 * median_length)].index
        outlier_values = df.loc[outlier_indices, alias].dropna().unique().tolist()
        if outlier_values:
            exceptions.append((alias, "Product Length Deviation Check", outlier_values))
    return exceptions

def customer_data_quality_checks(df, aliases):
    exceptions = []
    for alias in aliases:
        if "email" in alias.lower():
            invalid_emails = df[~df[alias].str.contains(r'^[\w\.-]+@[\w\.-]+\.\w+$', na=False, regex=True)][alias].dropna().unique().tolist()
            if invalid_emails:
                exceptions.append((alias, "Invalid Email Format Check", invalid_emails))
        if "phone" in alias.lower():
            phone_lengths = df[alias].astype(str).str.len()
            outlier_indices = phone_lengths[(phone_lengths < 7) | (phone_lengths > 15)].index
            outlier_values = df.loc[outlier_indices, alias].dropna().unique().tolist()
            if outlier_values:
                exceptions.append((alias, "Invalid Phone Length Check", outlier_values))
    return exceptions

def data_type_outlier_check(df, aliases):
    exceptions = []
    for alias in aliases:
        if df[alias].dtype == 'object':
            unexpected = df[alias].apply(lambda x: isinstance(x, (int, float)) if pd.notnull(x) else False)
        else:
            unexpected = df[alias].apply(lambda x: isinstance(x, str) if pd.notnull(x) else False)
        outlier_values = df.loc[unexpected, alias].dropna().unique().tolist()
        if outlier_values:
            exceptions.append((alias, "Data Type Outlier Check", outlier_values))
    return exceptions

# Data Quality Checks Loop
data_quality_exceptions = {}
for (table, domain), data in tables_and_attributes.items():
    attributes = data["attributes"]
    aliases = data["aliases"]

    if attributes:  # Proceed only if there are attributes
        select_clause = ", ".join([f"{attr} AS [{alias}]" for attr, alias in zip(attributes, aliases)])
        query = f"SELECT {select_clause} FROM {table}"

        try:
            df = pd.read_sql_query(query, conn)
            exceptions = []
            exceptions.extend(null_or_empty_check(df, aliases))
            exceptions.extend(frequent_category_check(df, aliases))
            for alias in aliases:
                exceptions.extend(length_consistency_check(df, alias, min_length=5, max_length=15))
            exceptions.extend(data_type_outlier_check(df, aliases))
            if domain == "Product":
                exceptions.extend(product_data_quality_checks(df, aliases))
            elif domain == "Customer":
                exceptions.extend(customer_data_quality_checks(df, aliases))

            if exceptions:
                if domain not in data_quality_exceptions:
                    data_quality_exceptions[domain] = []
                for alias, rule, values in exceptions:
                    data_quality_exceptions[domain].append((table, alias, rule, values))

        except Exception as e:
            print(f"Error performing checks for table {table} (Domain: {domain}): {e}")

# Generate exception reports for each domain
for domain, exceptions in data_quality_exceptions.items():
    wb = Workbook()
    ws = wb.active
    row = 1

    for table, alias, rule, values in exceptions:
        # Write table name, attribute, and rule
        ws.cell(row=row, column=1, value=table)
        ws.cell(row=row + 1, column=1, value=alias)
        ws.cell(row=row + 2, column=1, value=rule)

        # Write distinct exception values
        for i, value in enumerate(values, start=row + 3):
            ws.cell(row=i, column=1, value=value)

        # Move to the next block for the next attribute
        row += len(values) + 4

    # Save to an Excel file
    exception_report_path = f"C:/py/profiling_report/{domain}_data_quality_exceptions.xlsx"
    wb.save(exception_report_path)
    data_quality_exception_files.append(exception_report_path)
    print(f"Data quality exception report generated for domain: {domain}")

# Close the database connection
conn.close()

# Output the file paths
print("\nGenerated Profile Report HTML Files:")
for file in profile_report_html_files:
    print(file)

print("\nGenerated Data Quality Exception Excel Files:")
for file in data_quality_exception_files:
    print(file)
